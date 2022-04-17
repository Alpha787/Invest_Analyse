from multiprocessing.connection import wait
import requests
import time
import json
from colorama import init, Fore, Back, Style
from openpyxl import Workbook, load_workbook
import _init_
from datetime import datetime, timedelta, date
from get_t_info import (
    get_portfolio_response
)
import pandas as pd


init(autoreset=True)

headers = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'
# TODO: сделать поддержку api сервиса
# NOTE: почти сделана авторизация
# FIXME: доделать логин тинькофф

url = ("https://www.tinkoff.ru/login")

# def get_tinkoff_cookie() -> json: TODO: раскомментировать, если понадобится
#     # полезная функция для открытия окна браузера (ОСТАВИТЬ)
#     # wb.open_new(url) 
#     print(Fore.BLUE + "Открывается сайт для авторизации")
#     wb.get(using='safari').open_new(url)

#     # получаем куки
#     session = requests.Session()
#     r = session.get(url, headers={
#         "User-Agent":headers
#     })
#     cookie = requests.utils.dict_from_cookiejar(session.cookies)

#     if cookie:
#         time.sleep(5)

#     # запись куки в json, чтобы позже обновлять куки
#     with open('cookies.json', 'w') as f:
#         json.dump(requests.utils.dict_from_cookiejar(session.cookies), f)
#         print(Fore.GREEN + "Cookie получены!")
#     print(cookie)
        

#     # обновление куки для авторизации
#     with open('cookies.json') as f:
#         session.cookies.update(json.load(f))
#         print(Fore.GREEN + "Cookie обновлены!")
#     # r = requests.get(url, headers=headers)
#     # print(r.text)
# get_tinkoff_cookie()

# здесь должна быть автоматизация создания XLSX файла
def create_xlsx():
    wb = Workbook()
    get_data = _init_.Hola
    # get_data = get_portfolio_response()
    # wb = load_workbook('/Users/basil/Desktop/Project_Pycharm/tin_to_exel')
    filename = 'data_table_copy.xlsx'
    sheet_1 = wb.active
    sheet_1.title = "Книга 1"
    
    # cell = sheet_1['A7':'O7']
    starting_row = sheet_1['A7']
    starting_col = sheet_1['A8']

    # for item in get_data:
    # for row, tup in enumerate(get_data):
    #     for col, val in enumerate(tup):
            # sheet_1[row][col] = val
    #         print(row, col, val)
        # sheet_1.append(item)
        # print(item)
        # FIXME: возможно NonType возникает при записи 
        # в Openpyxl. Найти правильную функцию
        

    # for row in sheet_1['A17:O7']:
    #     sheet_1.append(row)

    # sheet_2 = wb.create_sheet(title="Информация по активам")
    # sheet_2['F5'] = 3.1415
    
    # sheet_3 = wb.create_sheet(title="Диаграммы")
    # for row in range(10, 20):
    #     for col in range(27, 54):
    #         _ = sheet_3.cell(column=col, row=row, value="{0}".format(col))

    # for sheet in book.worksheets:
    #     for row in get_data:
    #         sheet.append


    # print(ws3['AA10'].value)
    wb.save(filename=f'data_table_{date.today()}.xlsx')

create_xlsx()